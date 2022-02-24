from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference, Series

#  Criar workbook
pasta = Workbook()

# Renomear a planilha inicial e mudar cores
planilha_numeros = pasta.worksheets[0]
planilha_numeros.title = 'Numeros'
planilha_numeros.sheet_properties.tabColor = '8334EB'

# Criar planilha ByLearn
planilha_by_learn = pasta.create_sheet('ByLearn')
planilha_by_learn.sheet_properties.tabColor = '34EBB7'


# Criar outras Imagem
planilha_imagem = pasta.create_sheet('Imagem')
planilha_imagem.sheet_properties.tabColor = '80BF82'


# Criar planilha Gráficos
planilha_grafico = pasta.create_sheet('Gráficos')
planilha_grafico.sheet_properties.tabColor = 'CF9AFC'




# ----Planilha Numeros
# Inserir valores de 2 em 2

i=0
for linha in range(5):
    for coluna in range(10):
        celula = planilha_numeros.cell(row = linha + 1, column = coluna + 1, value = i)
        i = i+2

# ----Planilha ByLearn

# Cbeçalhos
planilha_by_learn['A1'] = 'Rede Social'
planilha_by_learn['B1'] = 'Link'

# Lista com os valores
lista = [
    ['Facebook', '@ByLearn'],
    ['Insta', '@ByLearn'],
    ['Youtube', '@ByLearn'],
    ['Dojo(Blog)','dojo.bylearn.com.br']
]

# Insere linhas na planilha
for linha in lista:
    planilha_by_learn.append(linha)


# ----Planilha Imagem

imagem = Image('c://teste/test_image.jpg')
planilha_imagem.add_image(imagem,'A1')


# ----Planilha Imagem
dados = [
    ('Prova', 'Competidor 1', 'Competidor 2'),
    ('Futebol', 10, 30),
    ('Basquete', 40, 60),
    ('Vôlei', 50, 70),
    ('Baseball', 20, 10),
    ('Corrida', 10, 40),
    ('Handball', 50, 30),
]

for linha in dados:
    planilha_grafico.append(linha)

grafico = BarChart()
grafico.title = 'Competição Esportiva'
grafico.y_axis.title = 'Pontuação'
grafico.x_axis.title = 'Prova'

# Configurar as referencias

pontuacao = Reference(planilha_grafico, min_col = 2, max_col = 3, min_row = 1, max_row = 7)
categoria_jogos = Reference(planilha_grafico, min_col = 1, max_col = 1, min_row = 2, max_row = 7)

# Configurar valores no grafico
grafico.add_data(pontuacao,titles_from_data = True)

# Configurar categorias do grafico
grafico.set_categories(categoria_jogos)

# Adicionar o Gráfico
planilha_grafico.add_chart(grafico)

pasta.save('Teste_arquivo_final.xlsx')